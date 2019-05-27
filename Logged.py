from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart,
    Reference
)
from openpyxl.styles import PatternFill
import os
import numpy

#Author: Ethan Newsom

#This program is intended to be used to grab pertinent information from logs in .xslx format generated by
#the INSPIRE game. The basic function of this program is as follows: It begins by taking the directory in
#which the logs are stored and iterates through the files. Once it finds a file, it grabs the name,
#increments the number of files found, and stores it into the specified worksheet. It then determines if
#the player finished the game.

#The script looks through the given sheet, looking for specified flags that indicate a change in conversation.
#Once it detects these flags, it then finds the choice numbers of each conversation point and saves them before
#moving back to finding the starts of the conversations.

#This script can be easily modified to add in further points of the game for data collection. Further
#documentation will be available so that others can work with this.

#Special acknowledgements go to Wookhee Min for helping me figure out my biggest roadblock and learning
#how to use the debugger and to Pengcheng Wang for helping me get Pycharm working.

# Activate a workbook to write into
wb = Workbook()
ws = wb.active

stdDevFinArray = []
stdDevAllArray = []

foundAthlete = False
foundMusician = False
foundSister = False
foundSisterAgain = False
foundVodka = False
filename = ''

rowSave = 1

#MAKE SOME PIE
def chartMaker(ws, wsload):
    pie = PieChart()
    labels = Reference(ws, min_col = 11, min_row = ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col = 12, min_row= ws.max_row - 4, max_row = ws.max_row)
    pie.add_data(data, titles_from_data= False)
    pie.set_categories(labels)
    pie.title = 'Athlete Choice 1'

    pie2 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=13, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie2.add_data(data, titles_from_data=False)
    pie2.set_categories(labels)
    pie2.title = 'Athlete Choice 2'

    pie3 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=14, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie3.add_data(data, titles_from_data=False)
    pie3.set_categories(labels)
    pie3.title = 'Athlete Texting'

    pie4 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=15, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie4.add_data(data, titles_from_data=False)
    pie4.set_categories(labels)
    pie4.title = 'Athlete Sister'

    pie5 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=16, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie5.add_data(data, titles_from_data=False)
    pie5.set_categories(labels)
    pie5.title = 'Musician Choice 1'

    pie6 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=17, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie6.add_data(data, titles_from_data=False)
    pie6.set_categories(labels)
    pie6.title = 'Musician Choice 2'

    pie7 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=18, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie7.add_data(data, titles_from_data=False)
    pie7.set_categories(labels)
    pie7.title = 'Musician Parents'

    pie8 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=19, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie8.add_data(data, titles_from_data=False)
    pie8.set_categories(labels)
    pie8.title = 'Sister Choice'

    pie9 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=20, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie9.add_data(data, titles_from_data=False)
    pie9.set_categories(labels)
    pie9.title = 'Vodka Choice'

    pie10 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=21, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie10.add_data(data, titles_from_data=False)
    pie10.set_categories(labels)
    pie10.title = 'Pong Choice'

    pie11 = PieChart()
    labels = Reference(ws, min_col=11, min_row=ws.max_row - 4, max_row=ws.max_row)
    data = Reference(ws, min_col=22, min_row=ws.max_row - 4, max_row=ws.max_row)
    pie11.add_data(data, titles_from_data=False)
    pie11.set_categories(labels)
    pie11.title = 'Slacker Choice'

    #Pie 1
    cell = ws.cell(row=ws.max_row + 2, column=1)

    ws.add_chart(pie, cell.coordinate)
    #Pie 2
    cell = ws.cell(row=ws.max_row, column=9)

    ws.add_chart(pie2, cell.coordinate)
    #Pie 3
    cell = ws.cell(row=ws.max_row, column=17)

    ws.add_chart(pie3, cell.coordinate)
    #Pie 4
    cell = ws.cell(row=ws.max_row + 15, column=1)

    ws.add_chart(pie4, cell.coordinate)
    #Pie 5
    cell = ws.cell(row=ws.max_row, column=9)

    ws.add_chart(pie5, cell.coordinate)
    #Pie 6
    cell = ws.cell(row=ws.max_row, column=17)

    ws.add_chart(pie6, cell.coordinate)
    #Pie 7
    cell = ws.cell(row=ws.max_row + 15, column=1)

    ws.add_chart(pie7, cell.coordinate)
    #Pie 8
    cell = ws.cell(row=ws.max_row, column=9)

    ws.add_chart(pie8, cell.coordinate)
    #Pie 9
    cell = ws.cell(row=ws.max_row, column=17)

    ws.add_chart(pie9, cell.coordinate)
    #Pie 10
    cell = ws.cell(row=ws.max_row + 15, column=1)

    ws.add_chart(pie10, cell.coordinate)
    #Pie 11
    cell = ws.cell(row=ws.max_row, column=9)

    ws.add_chart(pie11, cell.coordinate)

    return

#This function is meant to gather average times of each conversation, gameplay length, and finished time.
#It will also collect the number of times each choice was made for each character
def statGather(ws, cell):
    global rowSave
    global stdDevAllArray
    global stdDevFinArray

    rowCheck = 2
    completed = 0
    #This loop gathers the total number of successfully completed playthroughs.
    while rowCheck < rowSave + 1:
        cell = ws.cell(row = rowCheck, column = 3)
        if cell.value == 'Yes':
            completed += 1
        rowCheck += 1
    cell = ws.cell(row = rowSave + 1, column = 3)
    cell.value = completed

    #here is where we get the times
    col = 4
    while col < 12:
        #Iterate all the rows that have info in them.
        rowNums = 2
        meandiv = 0
        addedCell = ws.cell(row = rowSave + 1, column = col)
        addedCell.value = 0
        while rowNums < rowSave + 1:
            cell = ws.cell(row = rowNums, column = col)
            if cell.value != None and cell.value != 'DNF' and isinstance(cell.value, float):
                addedCell.value += cell.value
                meandiv += 1
            rowNums += 1
        #Gets the average time of all playthroughs.
        if meandiv != 0:
            addedCell.value = (addedCell.value/meandiv)
        col += 1

    #Adding the mean times label and the standard deviations for completed game times.
    cell = ws.cell(row = ws.max_row, column = 2)
    cell.value = 'Mean times:'
    cell = ws.cell(row=ws.max_row+1, column=2)
    cell.value = 'Standard Deviations'
    cell = ws.cell(row=ws.max_row, column = 3)
    cell.value = numpy.std(stdDevFinArray, axis = 0)
    cell = ws.cell(row=ws.max_row, column = 4)
    cell.value = numpy.std(stdDevFinArray, axis=0)


    #This part creates the table that counts the number of choices that are made
    cell = ws.cell(row = ws.max_row + 2, column = 11)
    cell.value = 'First Choice:'
    cell = ws.cell(row=ws.max_row + 1, column=11)
    cell.value = 'Second Choice:'
    cell = ws.cell(row=ws.max_row + 1, column=11)
    cell.value = 'Third Choice:'
    cell = ws.cell(row=ws.max_row + 1, column=11)
    cell.value = 'Fourth Choice:'
    cell = ws.cell(row=ws.max_row + 1, column=11)
    cell.value = 'No Response:'

    #This code gets the number of choices per question
    choiceCols = 12
    while choiceCols < 23:
        choiceRows = 2
        first = 0
        second = 0
        third = 0
        fourth = 0
        defaultChoice = 0
        #might have to change this up
        while choiceRows < ws.max_row - 6:
            cell = ws.cell(row = choiceRows, column = choiceCols)
            if cell.value == 0:
                first += 1
            if cell.value == 1:
                second += 1
            if cell.value == 2:
                third += 1
            if cell.value == 3:
                fourth += 1
            if cell.value == 'Default':
                defaultChoice += 1
            choiceRows += 1
        rowCheck = ws.max_row - 4
        while rowCheck < ws.max_row + 1:
            cell = ws.cell(row = rowCheck, column = choiceCols)
            if rowCheck == ws.max_row - 4:
                cell.value = first
            elif rowCheck == ws.max_row - 3:
                cell.value = second
            elif rowCheck == ws.max_row - 2:
                cell.value = third
            elif rowCheck == ws.max_row - 1:
                cell.value = fourth
            elif rowCheck == ws.max_row:
                cell.value = defaultChoice
            rowCheck += 1
        choiceCols += 1

    # This loop will highlight green any runs that completed the second half and yellow any runs of the first half.
    rowColorNum = 2
    fill = PatternFill("solid", fgColor="86EF41")
    while rowColorNum < ws.max_row - 7:
        cell = ws.cell(row=rowColorNum, column=3)
        if cell.value == 'Yes':
            colCount = 1
            while colCount < 23:
                cell = ws.cell(row=rowColorNum, column=colCount)
                cell.fill = fill
                colCount += 1
        rowColorNum += 1

    return
#this function gets the times at which a conversation is started based on the character mesh and
#whether the character has been spoken to yet.
#nested within the if statements will be the choicefinder function, so the rows that are iterated through will be returned
#in order to save time.
def timeFinder(wsLoad, cell, checkCell, row_count):
    global foundAthlete
    global foundMusician
    global foundSister
    global foundSisterAgain
    global foundVodka
    global rowSave

    #This checkCell is looking for the right mesh to indicate the start of a conversation.
    checkCell = wsLoad.cell(row=row_count, column=20)

    print (row_count)
    # First Athlete conversation
    if checkCell.value == 'Athlete_mesh' and not foundAthlete:
        checkCell = wsLoad.cell(row=row_count, column=6)
        cell = ws.cell(row=rowSave, column=6)
        cell.value = checkCell.value
        foundAthlete = True
        choiceCount = 4
        choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 12) #rowSave = 2 here
        """Note, this is just for experimental purposes, be sure to change above rowSave back """

    #First Musician Conversation
    elif checkCell.value == 'Creative_mesh' and not foundMusician:
        checkCell = wsLoad.cell(row=row_count, column=6)
        cell = ws.cell(row=rowSave, column=7)
        cell.value = checkCell.value
        foundMusician = True
        choiceCount = 3
        choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 16)

    #First LittleSister conversation, sends her upstairs automatically
    elif checkCell.value == 'LittleSister_mesh':
        if foundAthlete and not foundSister:
            checkCell = wsLoad.cell(row=row_count, column=6)
            cell = ws.cell(row=rowSave, column=8)
            cell.value = checkCell.value
            cell = ws.cell(row=rowSave, column=9)
            cell.value = 'None'
            foundSister = True
            foundSisterAgain = True
            choiceCount = 1
            choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 19)

        #Second LittleSister conversation
        elif foundAthlete and foundSister:
            checkCell = wsLoad.cell(row=row_count, column=6)
            cell = ws.cell(row=rowSave, column=9)
            cell.value = checkCell.value
            foundSister = True
            foundSisterAgain = True
            choiceCount = 1
            choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 19)

        #First LittleSister conversation, doesn't send her upstairs or have choices.
        elif not foundAthlete and not foundSister:
            checkCell = wsLoad.cell(row=row_count, column=6)
            cell = ws.cell(row=rowSave, column=8)
            cell.value = checkCell.value
            foundSister = True

    #The vodka conversation in all its glory.
    elif (checkCell.value == 'Creative_mesh') and foundAthlete and foundMusician and foundSisterAgain and not foundVodka:
        checkCell = wsLoad.cell(row=row_count, column=6)
        cell = ws.cell(row=rowSave, column=10)
        cell.value = checkCell.value
        choiceCount = 2
        foundVodka = True
        choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 20)

    elif (checkCell.value == 'Athlete_mesh') and foundAthlete and foundMusician and foundSisterAgain and not foundVodka:
        checkCell = wsLoad.cell(row=row_count, column=6)
        cell = ws.cell(row=rowSave, column=10)
        cell.value = checkCell.value
        choiceCount = 2
        foundVodka = True
        choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 20)

    #Slacker's mesh doesn't show since his dialogue occurs upon the opening of the door. Gotta grab his time when you first
    #see his name in the log after the vodka conversation
    checkCell = wsLoad.cell(row=row_count, column=14)
    if (checkCell.value == 'Outsider') and foundAthlete:
        checkCell = wsLoad.cell(row = row_count, column = 6)
        cell = ws.cell(row=rowSave, column=11)
        cell.value = checkCell.value
        choiceCount = 1
        choiceFinder(choiceCount, wsLoad, cell, checkCell, row_count, 22)

    return row_count

#choiceCount is the variable that tells how many choices are in the conversation.
#insert_col is the column that will contain the number of whichever choice is used.
#This function finds the choice number by checking for the "ChosenIndex" cell then getting the number one col over.
#The place it's inserted at can be incremented easily thanks to the setup of the spreadsheet and the dialogue flow
#of the game.

#insert_col is the first column that the values will be inserted into. The spreadsheet is set up so that we can iterate
#through and insert said values according to the order they are played.
def choiceFinder (choiceCount, wsLoad, cell, checkCell, row_count, insert_col):
    #count is here in order to jump over the log entries that have already been covered.

    global rowSave

    count = 0
    while choiceCount > 0 and row_count != wsLoad.max_row:
        checkCell = wsLoad.cell(row=row_count, column=21)
        count += 1
        if checkCell.value == 'ChosenIndex':
            checkCell = wsLoad.cell(row=row_count, column=22)
            cell = ws.cell(row=rowSave, column=insert_col)
            cell.value = checkCell.value
            insert_col += 1
            choiceCount -= 1
        row_count += 1
    return count


def main():
    # These are used to tell when these conversations have occured. Since after these, the conversations are more or less in
    # order. The checking will be done using the meshes in the log file (ie. Athlete_mesh) and if a play talks to that character
    # a second time, we don't want to be in a bad loop.
    """foundAthlete = False
    foundMusician = False
    foundSister = False
    foundSisterAgain = False
    foundVodka = False"""

    #for testing purposes
    global filename
    global rowSave
    global stdDevFinArray
    global stdDevAllArray

    # These are the numbers of choices in each conversation.

    print ('test text')

    # wb2 = load_workbook('test.xlsx')

    ws['A1'] = 'Filename'
    ws['B1'] = 'Log File #'
    ws['C1'] = 'Finished'
    ws['D1'] = 'Gameplay Duration'
    ws['E1'] = 'Finished Time'
    ws['F1'] = 'Athlete Timestamp'
    ws['G1'] = 'Musician Timestamp'
    ws['H1'] = 'Sister Timestamp'
    ws['I1'] = 'Second Sister Timestamp'
    ws['J1'] = 'Vodka Timestamp'
    ws['K1'] = 'Slacker Timestamp'
    ws['L1'] = 'Athlete Choice1'
    ws['M1'] = 'Athlete Choice2'
    ws['N1'] = 'AthTexting Choice'
    ws['O1'] = 'AthleteSister Choice'
    ws['P1'] = 'Musician Choice1'
    ws['Q1'] = 'Musician Choice2'
    ws['R1'] = 'MusicParent Choice'
    ws['S1'] = 'Sister Choice'
    ws['T1'] = 'Vodka Choice'
    ws['U1'] = 'Pong Choice'
    ws['V1'] = 'Slacker Choice'

    # The number of logs that we are dealing with. Used to help denote the number of rows we're dealing with and, of course,
    # the number of logs we've gone through.
    logNumber = 0

    #rowSave is the row that is currently being worked on.
    rowSave = 1

    # cell represents the cells in the worksheet being written to.
    cell = ws.cell(row=1, column=1)

    # Gotta get the directory from the user here.
    directory = ''
    while not os.path.isdir(directory):
        print 'Example: /Users/cenewsom/Documents/OctStudyT51T52/TestNew'
        directory = raw_input('Please enter the full directory name: ')

    #directory = '/Users/cenewsom/Documents/StudyMarch2016/LogData/TestFolder'  # obviously will need to change. Will make some change in the code so that input is accepted

    for filename in os.listdir(directory):
        global foundAthlete
        global foundMusician
        global foundSister
        global foundSisterAgain
        global foundVodka

        print(filename)
        if filename.endswith(".xlsx"):

            wbLoad = load_workbook(directory+'/'+filename)
            wsLoad = wbLoad.get_sheet_by_name('Worksheet') #print sheet otherwise
            logNumber += 1
            colSave = 1
            cell = ws.cell(row=(rowSave+1), column=colSave)

            #Current working directory
            cell.value = filename

            #This puts the log number down
            colSave = 2
            cell = ws.cell(row=(rowSave+1), column=colSave)
            cell.value = logNumber
            rowSave += 1

            #End time of the log
            checkCell = wsLoad.cell(row=wsLoad.max_row, column=4)

            row_count = 1

            #All reading occurs in here.
            while row_count != wsLoad.max_row:
                timeFinder(wsLoad, cell, checkCell, row_count)
                row_count += 1

            checkCell = wsLoad.cell(row=wsLoad.max_row, column=4)

            #Game ends after getting every conversation and ending correctly
            if checkCell.value == 'Ended' and foundVodka:
                # marks down the time the game ended, regardless of completion
                cell = ws.cell(row=rowSave, column=4)
                checkCell = wsLoad.cell(row=wsLoad.max_row, column=6)
                cell.value = checkCell.value
                # Marks down the finished time if the game was completed
                cell = ws.cell(row=rowSave, column=5)
                checkCell = wsLoad.cell(row=wsLoad.max_row, column=6)
                cell.value = checkCell.value
                stdDevFinArray.append(cell.value)
                stdDevAllArray.append(cell.value)
                # marks whether the user completed the game or not
                cell = ws.cell(row=rowSave, column=3)
                cell.value = 'Yes'
            #ends otherwise
            else:
                cell = ws.cell(row=rowSave, column=4)
                checkCell = wsLoad.cell(row=wsLoad.max_row, column=6)
                cell.value = checkCell.value
                stdDevAllArray.append(cell.value)
                cell = ws.cell(row=rowSave, column=5)
                cell.value = 'DNF'
                cell = ws.cell(row=rowSave, column=3)
                cell.value = 'No'

            #Reset bool values
            foundAthlete = False
            foundMusician = False
            foundSister = False
            foundSisterAgain = False
            foundVodka = False

    statGather(ws, cell)

    chartMaker(ws, wsLoad)

    wb.save('March2016Logged.xlsx')
    print 'The dastardly deed you requested has been done.'

if __name__ == '__main__':
    main()